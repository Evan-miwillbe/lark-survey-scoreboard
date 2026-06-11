"""
Import Excel survey data into Feishu Base.
Reads the Excel file and creates records via Feishu Open API.
Deduplication: same name+phone → overwrites with latest.
"""
import sys, os, json, time, re
from openpyxl import load_workbook

# --- Config ---
EXCEL_PATH = r"D:/others/微信/xwechat_files/wxid_rgtx7gtib1ny22_598c/msg/file/2026-06/诚信销售作业二汇总(1).xlsx"
SURVEY_CONFIG_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "src", "survey-config.json")

BASE_TOKEN = "STIKbVfxvaxd4PsvdBKcaQTKnFe"
TABLE_ID = "tbl2uUjvQ4MochWh"
API_BASE = "https://open.feishu.cn"

# Credentials from lark-cli
lark_config_path = os.path.expanduser(r"~\.lark-cli\config.json")
lark_secret_path = os.path.expanduser(r"~\.lark-cli\appsecret.txt")
with open(lark_config_path) as f:
    lark_config = json.load(f)
APP_ID = lark_config["apps"][0]["appId"]
with open(lark_secret_path) as f:
    APP_SECRET = f.read().strip()

# Identity field IDs in Feishu Base
ID_FIELDS = {
    "您的姓名": "fldMYZxXcY",
    "手机号": "fldcKpo8CT",
    "部门名称": "fld234JRZA",
}

# --- Load survey config ---
with open(SURVEY_CONFIG_PATH, "r", encoding="utf-8") as f:
    survey_config = json.load(f)

# Build question index -> field_id map (0-based index)
q_field_map = {}
for q in survey_config["questions"]:
    q_field_map[q["number"] - 1] = q["id"]

def strip_question_prefix(text):
    """Remove leading number prefix like '2、' or '120、' from question text."""
    return re.sub(r'^\d+[、.]\s*', '', str(text or '')).strip()

def get_tenant_token():
    url = f"{API_BASE}/open-apis/auth/v3/tenant_access_token/internal"
    resp = __import__('urllib.request').request.urlopen(
        __import__('urllib.request').request.Request(
            url,
            data=json.dumps({"app_id": APP_ID, "app_secret": APP_SECRET}).encode(),
            headers={"Content-Type": "application/json; charset=utf-8"}
        )
    )
    data = json.loads(resp.read())
    if data.get("code") != 0:
        raise Exception(f"Token failed: {data}")
    return data["tenant_access_token"]

def api_request(token, method, url_path, body=None):
    import urllib.request as ur
    full_url = url_path if url_path.startswith("http") else f"{API_BASE}{url_path}"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json; charset=utf-8"
    }
    data_bytes = json.dumps(body).encode() if body else None
    req = ur.Request(full_url, data=data_bytes, headers=headers, method=method)
    resp = ur.urlopen(req)
    return json.loads(resp.read())

def create_records(token, records_data):
    """Batch create records (max 500 per batch)."""
    url = f"/open-apis/bitable/v1/apps/{BASE_TOKEN}/tables/{TABLE_ID}/records"
    return api_request(token, "POST", url, {"records": records_data})

def delete_all_records(token):
    """Delete all existing records to avoid duplicates."""
    all_ids = []
    page_token = None
    while True:
        url = f"/open-apis/bitable/v1/apps/{BASE_TOKEN}/tables/{TABLE_ID}/records?page_size=500"
        if page_token:
            url += f"&page_token={page_token}"
        result = api_request(token, "GET", url)
        items = result.get("data", {}).get("items") or []
        all_ids.extend([item["record_id"] for item in items])
        if not result.get("data", {}).get("has_more"):
            break
        page_token = result["data"]["page_token"]

    print(f"  Found {len(all_ids)} existing records to delete")
    # Delete in batches (max 500)
    for i in range(0, len(all_ids), 500):
        batch = all_ids[i:i+500]
        # Need to delete one by one or use batch delete
        for rid in batch:
            try:
                api_request(token, "DELETE",
                    f"/open-apis/bitable/v1/apps/{BASE_TOKEN}/tables/{TABLE_ID}/records/{rid}")
            except Exception as e:
                print(f"  Delete warning: {e}")
    return len(all_ids)

def main():
    print("=== Import Excel Data to Feishu Base ===")

    # Read Excel
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["Sheet1"]

    # Header: row 1, columns 1-11
    header = [ws.cell(row=1, column=c).value for c in range(1, 12)]
    person_names = [str(ws.cell(row=1, column=c).value).strip() for c in range(3, 12)]
    print(f"People: {person_names}")

    # Build records
    records = []
    for person_idx, name in enumerate(person_names):
        fields = {
            ID_FIELDS["您的姓名"]: name,
            ID_FIELDS["手机号"]: f"1380000{person_idx:04d}",
            ID_FIELDS["部门名称"]: "销售部",
        }

        for row in range(2, 122):  # rows 2-121 are questions 1-120
            score_raw = ws.cell(row=row, column=person_idx + 3).value
            try:
                score_val = int(float(score_raw))
            except (ValueError, TypeError):
                continue

            q_idx = row - 2  # 0-based question index
            field_id = q_field_map.get(q_idx)
            if field_id:
                fields[field_id] = score_val

        filled = len(fields) - 3
        records.append({"fields": fields})
        print(f"  {name}: {filled}/120 scores mapped")

    # Get token and import
    print("\nAuthenticating with Feishu...")
    token = get_tenant_token()
    print("Token obtained.")

    # Delete existing records
    print("\nClearing existing records...")
    deleted = delete_all_records(token)
    print(f"Deleted {deleted} existing records.")

    # Create new records
    print(f"\nCreating {len(records)} records...")
    result = create_records(token, records)

    if result.get("code") == 0:
        created = len(result.get("data", {}).get("records", []))
        print(f"SUCCESS: Created {created} records!")
    else:
        print(f"ERROR: {result}")
        sys.exit(1)

    # Verify
    print("\nVerifying...")
    verify = api_request(token, "GET",
        f"/open-apis/bitable/v1/apps/{BASE_TOKEN}/tables/{TABLE_ID}/records?page_size=5")
    count = len(verify.get("data", {}).get("items", []))
    has_more = verify.get("data", {}).get("has_more", False)
    total = verify.get("data", {}).get("total", 0)
    print(f"Records in Base: {total} (has_more={has_more})")
    print("\nDone! Dashboard should now show data at:")
    print("  https://evan-miwillbe.github.io/lark-survey-scoreboard/dashboard.html")

if __name__ == "__main__":
    main()
