"""
Import Excel survey data into Feishu Base using lark-cli.
One record per person with all 120 scores. Uses field IDs for compactness.
"""
import subprocess, json, os, sys
from openpyxl import load_workbook

EXCEL_PATH = r"D:/others/微信/xwechat_files/wxid_rgtx7gtib1ny22_598c/msg/file/2026-06/诚信销售作业二汇总(1).xlsx"
SURVEY_CONFIG_PATH = r"C:\Users\Tengm\lark-survey-scoreboard\src\survey-config.json"
BASE_TOKEN = "STIKbVfxvaxd4PsvdBKcaQTKnFe"
TABLE_ID = "tbl2uUjvQ4MochWh"
LARK_CLI = r"C:\Users\Tengm\AppData\Roaming\npm\lark-cli.cmd"

with open(SURVEY_CONFIG_PATH, "r", encoding="utf-8") as f:
    survey_config = json.load(f)

# Build field ID map (0-based index -> field_id)
q_field_map = {}
for q in survey_config["questions"]:
    q_field_map[q["number"] - 1] = q["id"]

wb = load_workbook(EXCEL_PATH, data_only=True)
ws = wb["Sheet1"]

# Column 3-11 are people (columns C-K in Excel, 1-based)
person_names = [str(ws.cell(row=1, column=c).value or "").strip() for c in range(3, 12)]
print(f"People: {person_names}")

records_created = 0
for person_idx, name in enumerate(person_names):
    fields = {
        "fldMYZxXcY": name,
        "fldcKpo8CT": f"1380000{person_idx:04d}",
        "fld234JRZA": "销售部",
    }
    for row in range(2, 122):
        score_raw = ws.cell(row=row, column=person_idx + 3).value
        try:
            score_val = int(float(score_raw))
        except (ValueError, TypeError):
            continue
        field_id = q_field_map.get(row - 2)
        if field_id:
            fields[field_id] = score_val

    json_str = json.dumps(fields, ensure_ascii=False)
    print(f"  {name}: {len(fields)-3} scores, JSON={len(json_str)} chars...")

    try:
        result = subprocess.run(
            [LARK_CLI, "base", "+record-upsert",
             "--base-token", BASE_TOKEN,
             "--table-id", TABLE_ID,
             "--json", json_str],
            capture_output=True, text=True, timeout=60
        )
        data = json.loads(result.stdout)
        if data.get("ok"):
            rid = data.get("data", {}).get("record", {}).get("record_id_list", ["?"])[0]
            print(f"    OK → {rid}")
            records_created += 1
        else:
            err = data.get("error", {})
            print(f"    FAIL: {err.get('message', result.stdout[:300])}")
            # If the JSON is too long, try writing to file and using stdin
            if "too long" in str(err).lower() or len(json_str) > 7000:
                print(f"    Retrying with file-based approach...")
                tmp_path = os.path.join(os.environ.get("TEMP", "/tmp"), f"lark_import_{person_idx}.json")
                with open(tmp_path, "w", encoding="utf-8") as f:
                    f.write(json_str)
                # Use PowerShell to read file and pass to lark-cli
                ps_cmd = f'$json = Get-Content -Raw "{tmp_path}"; & "{LARK_CLI}" base +record-upsert --base-token {BASE_TOKEN} --table-id {TABLE_ID} --json $json'
                result2 = subprocess.run(["powershell", "-NoProfile", "-Command", ps_cmd],
                    capture_output=True, text=True, timeout=60)
                print(f"    PS result: {result2.stdout[:300]}")
                os.unlink(tmp_path)
    except Exception as e:
        print(f"    ERROR: {e}")

print(f"\nDone: {records_created}/{len(person_names)} records created")
